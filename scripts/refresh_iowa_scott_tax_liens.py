#!/usr/bin/env python3
"""Refresh Scott County, Iowa 2026 property-level tax-sale liens.

Scott County's own annual sale was held 2026-06-15; this collector reads the
newest post-sale list the county publishes (2026-06-16 as of this writing --
dated the day after the sale, and the county has not posted a newer file
since). Unlike the pre-sale lists, this one is the set of items that did NOT
sell to a public bidder and are now county-held certificates, available for
assignment from the Treasurer's office rather than a future public auction.

The workbook is a running per-parcel receipt ledger, not a flat one-row-per-
item list: a single Item Number can carry a "CT" (certificate) or "MH"
(mobile home) row plus several "SA" (special assessment) sibling rows for the
same item/year, and -- for parcels that went unsold across more than one
annual sale -- older "DT" rows from prior years. Only the CT/MH row for each
item is published; SA/DT rows are never emitted as separate records.

Empirically verified (2026-08-25, 64 sample items, 100% match, and confirmed
textually against the county's own 2026 tax sale rules, which state the
tax-sale certificate fee is $20.00): the CT/MH row's own "Sale Amount" field
already equals the sum of (First Half + Second Half + Interest + Costs) across
that row and every SA sibling row for the same item, plus the flat $20.00
statutory certificate fee. This collector re-derives that sum for every row it
publishes and refuses to publish any item where the two don't match, rather
than trusting the source's own the "Sale Amount" figure blindly.

The source publishes real owner/taxpayer names and mailing addresses (`Name
1/2/3`, `Address Attention/Lines/City State/Zip 1/2/3`). Those columns are
never read into a variable anywhere in this file.
"""

from __future__ import annotations

import io
import json
from datetime import date, datetime, timezone
from pathlib import Path

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DETAILS = ROOT / "data" / "tax-lien-properties.json"
PROFILE_ID = "IA-Scott-2026"
SOURCE_PAGE = "https://www.scottcountyiowa.gov/treasurer/tax-sale"
SOURCE_XLSX = "https://www3.scottcountyiowa.gov/treasurer/pub/tax_sale/2026/20260616_Tax_Sale_List.xlsx"
RULES_PDF = "https://www3.scottcountyiowa.gov/treasurer/pub/tax_sale_rules/2026_Annual_Tax_Sale_Rules.pdf"
UA = "TaxLienGuideBot/2.9 (public tax-lien research; no access-control bypass)"

CERTIFICATE_FEE = 20.00
FEE_TOLERANCE = 0.01
PRIMARY_TYPES = {"CT", "MH"}
MIN_EXPECTED_ROWS = 150

# Columns intentionally never read: any "Name"/"Address"/"Zip" column belongs
# to the delinquent taxpayer's mailing record, not the property, and this
# project never collects owner/taxpayer identity (see BUG-004/BUG-005).
COMPONENT_FIELDS = ("First Half", "Second Half", "Interest", "Costs")


def fetch_xlsx() -> bytes:
    response = requests.get(SOURCE_XLSX, headers={"User-Agent": UA, "Accept": "*/*"}, timeout=90)
    response.raise_for_status()
    return response.content


def _columns(header: tuple) -> dict[str, int]:
    return {str(name): i for i, name in enumerate(header) if name}


def _num(value) -> float:
    return float(value) if isinstance(value, (int, float)) else 0.0


def parse_rows(all_rows: list[tuple], verified: str, min_expected: int = MIN_EXPECTED_ROWS) -> list[dict]:
    if not all_rows:
        raise RuntimeError("Scott County workbook is empty")
    columns = _columns(all_rows[0])
    required = {"Item Number", "Year", "Type", "Parcel", "Situs Address", "Legal Description",
                "Net Acres", "Class", "Sale Amount", *COMPONENT_FIELDS}
    missing = required - columns.keys()
    if missing:
        raise RuntimeError(f"Scott County workbook is missing expected columns: {sorted(missing)}")

    by_item: dict[int, list[tuple]] = {}
    for record in all_rows[1:]:
        item_number = record[columns["Item Number"]]
        if not isinstance(item_number, int):
            continue
        by_item.setdefault(item_number, []).append(record)

    rows: list[dict] = []
    skipped_missing_amount = 0
    skipped_fee_mismatch = 0
    skipped_ambiguous = 0

    for item_number, records in sorted(by_item.items()):
        primaries = [r for r in records if r[columns["Type"]] in PRIMARY_TYPES]
        if len(primaries) != 1:
            skipped_ambiguous += 1
            continue
        primary = primaries[0]
        sale_amount = primary[columns["Sale Amount"]]
        if not isinstance(sale_amount, (int, float)):
            skipped_missing_amount += 1
            continue

        component_total = sum(_num(r[columns[f]]) for r in records for f in COMPONENT_FIELDS)
        if abs(float(sale_amount) - component_total - CERTIFICATE_FEE) > FEE_TOLERANCE:
            skipped_fee_mismatch += 1
            continue

        is_mobile_home = primary[columns["Type"]] == "MH"
        parcel_class = primary[columns["Class"]]
        if is_mobile_home:
            property_type = "Mobile home"
        elif parcel_class == "C":
            property_type = "Real estate (commercial class)"
        else:
            property_type = "Real estate"

        acreage = primary[columns["Net Acres"]]
        rows.append({
            "record_id": f"IA-Scott-2026-{item_number}",
            "profile_id": PROFILE_ID,
            "state": "IA",
            "state_name": "Iowa",
            "county": "Scott",
            "parcel_id": primary[columns["Parcel"]],
            "sale_item_number": str(item_number),
            "property_address": primary[columns["Situs Address"]],
            "address": primary[columns["Situs Address"]],
            "city": None,
            "zip": None,
            "legal_description": primary[columns["Legal Description"]],
            "property_type": property_type,
            "acreage": acreage if isinstance(acreage, (int, float)) and acreage > 0 else None,
            "auction_date": None,
            "sale_date": "2026-06-15",
            "auction_time": None,
            "auction_format": "County-held certificate; did not sell to a public bidder at the 2026-06-15 "
                               "annual sale and is now available for assignment from the Treasurer's office, "
                               "not a scheduled future public auction",
            "auction_location": "Scott County Treasurer's Office, 600 W. 4th Street, Davenport, IA 52801",
            "auction_url": SOURCE_PAGE,
            "official_source_url": SOURCE_XLSX,
            "direct_listing_url": SOURCE_PAGE,
            "minimum_bid": None,
            "opening_bid": None,
            "delinquent_tax_amount": float(sale_amount),
            "fees_costs": f"Includes the county's statutory ${CERTIFICATE_FEE:.2f} tax-sale certificate fee; "
                           "special-assessment collection fees are embedded in the total, not itemized per row",
            "assessed_value": None, "market_value": None,
            "tax_years_delinquent": None,
            "sale_status": "County-held certificate snapshot as of the county's 2026-06-16 post-sale publication "
                            "(the newest list the county has published as of this refresh); status must be "
                            "reconfirmed with the Treasurer before purchase",
            "lien_type": "Iowa tax sale certificate / property-tax lien (county-held)",
            "sale_type": "tax_lien",
            "maximum_statutory_return": "2% per month redemption interest; fractions of a month count as a whole month",
            "winning_rate_mechanism": "Not a live bid -- this certificate is assigned by the County Treasurer to the "
                                       "requesting purchaser for the full published amount plus a $100 assignment fee",
            "redemption_period": "A county-held certificate gives the assignee only 3 years from the date of "
                                  "assignment to qualify for a deed (shorter than a certificate purchased at the "
                                  "live sale); verify with the Treasurer before purchase",
            "important_rules": "This is a county-held certificate from an item that received no bidder at the "
                                "live annual sale, not a preview of an upcoming auction. Contact the Treasurer's "
                                "office to assign a certificate; the assignment fee is $100. Owner/taxpayer names "
                                "published alongside this row in the source are intentionally not collected.",
            "data_source": "Scott County Treasurer official 2026 post-sale tax sale list (XLSX)",
            "last_verified": verified,
            "source_mode": "official_post_sale_snapshot",
            "data_completeness": {"published_fields": 8, "tracked_fields": 19, "percent": 42},
            "research_priority": {
                "score": 38,
                "label": "Low research priority",
                "reasons": [
                    "Official parcel ID is published",
                    "A situs address is available for follow-up research",
                    "The source publishes a legal description",
                    "The county-verified current amount owed is published",
                ],
                "disclaimer": "Research-priority ranking only; not a buy recommendation.",
            },
        })

    if len(rows) < min_expected:
        raise RuntimeError(
            f"Scott County parser produced only {len(rows)} rows (skipped: "
            f"{skipped_missing_amount} missing Sale Amount, {skipped_fee_mismatch} failed the fee invariant, "
            f"{skipped_ambiguous} had zero/multiple primary rows); expected at least {min_expected}"
        )
    forbidden_keys = {"owner", "owner_name", "taxpayer", "taxpayer_name", "mailing_address"}
    if any(forbidden_keys.intersection(row) for row in rows):
        raise RuntimeError("Scott County output contains a restricted owner/taxpayer-name field")
    return rows


def existing_rows() -> list[dict]:
    if not DETAILS.exists():
        return []
    doc = json.loads(DETAILS.read_text(encoding="utf-8"))
    return [row for row in doc.get("properties", []) if row.get("profile_id") == PROFILE_ID]


def update_details(rows: list[dict]) -> None:
    doc = json.loads(DETAILS.read_text(encoding="utf-8"))
    profiles = doc.setdefault("profiles", {})
    profiles[PROFILE_ID] = {
        "state": "IA", "state_name": "Iowa", "county": "Scott",
        "sale_date": "2026-06-15",
        "auction_format": "County-held certificates available by assignment from the Treasurer",
        "auction_location": "Scott County Treasurer's Office, 600 W. 4th Street, Davenport, IA 52801",
        "auction_url": SOURCE_PAGE,
        "direct_listing_url": SOURCE_PAGE,
        "official_source_url": SOURCE_XLSX,
        "lien_type": "Iowa tax sale certificate / property-tax lien (county-held)",
        "sale_type": "tax_lien",
        "maximum_statutory_return": "2% per month redemption interest; fractions of a month count as a whole month",
        "winning_rate_mechanism": "County-Treasurer assignment of an unsold certificate, not a live bid",
        "important_rules": "County-held certificates from items that received no bidder at the 2026-06-15 annual "
                            "sale. Assignment fee is $100, separate from the published certificate amount. Owner "
                            "names in the source are intentionally not collected.",
        "data_source": "Scott County Treasurer official 2026 post-sale tax sale list (XLSX)",
        "last_verified": date.today().isoformat(),
        "source_mode": "official_post_sale_snapshot",
        "county_information_url": SOURCE_PAGE,
        "procedures_url": RULES_PDF,
    }
    doc["properties"] = [row for row in doc.get("properties", []) if row.get("profile_id") != PROFILE_ID] + rows
    doc["updated_at"] = datetime.now(timezone.utc).isoformat()
    full = [{**profiles.get(row.get("profile_id"), {}), **row} for row in doc["properties"]]
    doc["counts"] = {
        "total_records": len(full),
        "states": len({row.get("state") for row in full if row.get("state")}),
        "counties": len({(row.get("state"), row.get("county")) for row in full if row.get("state") and row.get("county")}),
        "with_parcel_id": sum(bool(row.get("parcel_id")) for row in full),
        "with_address": sum(bool(row.get("property_address")) for row in full),
        "with_auction_date": sum(bool(row.get("auction_date")) for row in full),
        "with_minimum_bid": sum(row.get("minimum_bid") is not None for row in full),
        "with_assessed_value": sum((row.get("assessed_value") is not None or row.get("market_value") is not None) for row in full),
        "with_research_priority": sum(bool(row.get("research_priority")) for row in full),
    }
    DETAILS.write_text(json.dumps(doc, separators=(",", ":")), encoding="utf-8")


def main() -> None:
    try:
        raw = fetch_xlsx()
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        all_rows = list(sheet.iter_rows(values_only=True))
        rows = parse_rows(all_rows, date.today().isoformat())
    except (requests.RequestException, RuntimeError) as exc:
        prior = existing_rows()
        if not prior:
            raise
        print(f"Scott County IA: source unavailable/unparseable; preserved {len(prior)} previously verified rows. Reason: {exc}")
        return
    update_details(rows)
    print(f"Scott County IA: loaded {len(rows)} county-held certificate rows; taxpayer names intentionally omitted")


if __name__ == "__main__":
    main()
