#!/usr/bin/env python3
"""Build the property-level tax-lien dataset from public official lists.

The collector intentionally excludes owner names.  A failed source refresh never
deletes the last verified rows for that source; the prior rows are retained and
the failure is recorded in ``source_health``.
"""

from __future__ import annotations

import io
import json
import re
import ssl
import sys
import urllib.error
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data" / "tax-lien-properties.json"
LEGACY_PROPERTIES = ROOT / "data" / "properties.json"
USER_AGENT = "tax-lien-guide-public-data-refresh/1.0 (+https://juliaonmoon.github.io/tax-lien-guide/)"

ALLEN_PAGE = "https://www.allencounty.in.gov/270/Tax-Sale"
ALLEN_XLSX = "https://www.allencounty.in.gov/DocumentCenter/View/14707/News-A-Z-2026-final-81426"
ALLEN_AUCTION = "https://www.govease.com/auctions"
TIPPECANOE_PAGE = "https://www.tippecanoe.in.gov/199/Tax-Sale"
TIPPECANOE_PDF = "https://www.tippecanoe.in.gov/DocumentCenter/View/53944"
TIPPECANOE_LISTING = (
    "https://properties.sriservices.com/properties?county=Tippecanoe&saleDate=AF&"
    "saleDateVal=&saleFormat=&saleStatus=&saleType=F&searchText=&sortBy=AD&"
    "sortOrder=desc&state=IN"
)
WABASH_PAGE = "https://www.in.gov/counties/wabash/"
WABASH_PDF = "https://www.in.gov/counties/wabash/files/legal-notices/Wabash-County-2026-Tax-Sale-Ad.pdf"
WABASH_LISTING = (
    "https://properties.sriservices.com/properties?county=Wabash&saleDate=AF&"
    "saleDateVal=&saleFormat=&saleStatus=&saleType=F&searchText=&sortBy=AD&"
    "sortOrder=desc&state=IN"
)
GRANT_PAGE = "https://www.grantcounty.in.gov"
GRANT_PDF = "https://www.in.gov/counties/grant/files/Grant-County-2026-Certificate-Sale-Advertisement.pdf"
GRANT_LISTING = (
    "https://properties.sriservices.com/properties?county=Grant&saleDate=AF&"
    "saleDateVal=&saleFormat=&saleStatus=&saleType=F&searchText=&sortBy=AD&"
    "sortOrder=desc&state=IN"
)
HAMILTON_PAGE = "https://www.hamiltoncounty.in.gov/452/Real-Property-Tax-Sale"
HAMILTON_LISTING_APP = "https://secure2.hamiltoncounty.in.gov/taxsale/"
HAMILTON_API = "https://secure2.hamiltoncounty.in.gov/TaxSale/TaxSaleWeb/TaxSale_Read"
HAMILTON_LISTING = (
    "https://properties.sriservices.com/properties?county=Hamilton&saleDate=AF&"
    "saleDateVal=&saleFormat=&saleStatus=&saleType=F&searchText=&sortBy=AD&"
    "sortOrder=desc&state=IN"
)

REQUIRED_FIELDS = [
    "state", "county", "parcel_id", "sale_item_number", "property_address",
    "city", "zip", "legal_description", "property_type", "auction_date",
    "auction_time", "auction_format", "auction_location", "auction_url",
    "official_source_url", "minimum_bid", "delinquent_tax_amount", "fees_costs",
    "assessed_value", "acreage", "tax_years_delinquent", "sale_status",
    "lien_type", "maximum_statutory_return", "winning_rate_mechanism",
    "redemption_period", "important_rules", "data_source", "last_verified",
    "data_completeness", "research_priority",
]


def fetch(url: str, timeout: int = 90) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT, "Accept": "*/*"})
    context = ssl.create_default_context()
    with urllib.request.urlopen(req, timeout=timeout, context=context) as response:
        return response.read()


def clean(value) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def number(value) -> float | None:
    if value in (None, ""):
        return None
    try:
        return round(float(str(value).replace("$", "").replace(",", "")), 2)
    except (TypeError, ValueError):
        return None


def completeness(row: dict) -> dict:
    useful = [
        "parcel_id", "sale_item_number", "property_address", "city", "zip",
        "legal_description", "property_type", "auction_date", "auction_time",
        "auction_format", "auction_url", "minimum_bid", "delinquent_tax_amount",
        "fees_costs", "assessed_value", "acreage", "tax_years_delinquent",
        "sale_status", "official_source_url",
    ]
    published = sum(row.get(key) not in (None, "", []) for key in useful)
    pct = round(published / len(useful) * 100)
    return {"published_fields": published, "tracked_fields": len(useful), "percent": pct}


def research_priority(row: dict) -> dict | None:
    """Rank research usefulness only when four or more concrete signals exist."""
    reasons = []
    score = 0
    if row.get("parcel_id"):
        score += 12
        reasons.append("Official parcel ID is published")
    if row.get("property_address"):
        score += 16
        reasons.append("A situs address is available for follow-up research")
    if row.get("legal_description"):
        score += 10
        reasons.append("The source publishes a legal description")
    if row.get("minimum_bid") is not None:
        score += 12
        reasons.append("A current minimum bid is published")
    if row.get("assessed_value") is not None:
        score += 12
        reasons.append("Official assessed/appraised value is available")
    if row.get("auction_date"):
        score += 8
        reasons.append("The auction date is confirmed")
    if row.get("auction_url"):
        score += 5
        reasons.append("A direct auction or listing link is available")
    if len(reasons) < 4:
        return None
    score = min(score, 100)
    label = "High" if score >= 70 else "Medium" if score >= 45 else "Low"
    return {
        "score": score,
        "label": f"{label} research priority",
        "reasons": reasons,
        "disclaimer": "Research-priority ranking only; not a buy recommendation.",
    }


def finish(row: dict) -> dict:
    for field in REQUIRED_FIELDS:
        row.setdefault(field, None)
    row["data_completeness"] = completeness(row)
    row["research_priority"] = research_priority(row)
    return row


def allen_rows(raw: bytes, verified: str) -> list[dict]:
    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet = workbook.active
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        item, parcel = clean(values[0]), clean(values[1])
        if not (item and item.isdigit() and parcel and re.fullmatch(r"\d{2}-\d{2}-\d{2}-\d{3}-\d{3}\.\d{3}-\d{3}", parcel)):
            continue
        address = clean(values[9])
        city = clean(values[12])
        zipcode = clean(values[13])
        row = {
            "record_id": f"IN-Allen-2026-{item}",
            "state": "IN", "state_name": "Indiana", "county": "Allen",
            "parcel_id": parcel, "sale_item_number": item,
            "certificate_number": None, "property_address": address,
            "address": address, "city": city, "zip": zipcode,
            "legal_description": clean(values[8]), "property_type": None,
            "auction_date": "2026-09-16", "sale_date": "2026-09-16",
            "auction_time": "09:00 ET", "auction_format": "Online",
            "auction_location": "GovEase online auction", "auction_url": ALLEN_AUCTION,
            "direct_listing_url": ALLEN_PAGE, "official_source_url": ALLEN_XLSX,
            "minimum_bid": number(values[16]), "opening_bid": number(values[16]),
            "delinquent_tax_amount": None,
            "fees_costs": "$620 tax-sale costs are included in the minimum bid; not separated per row",
            "assessed_value": None, "market_value": None, "acreage": None,
            "tax_years_delinquent": None, "sale_status": "Published — subject to removal or change",
            "lien_type": "Indiana tax sale certificate / lien", "sale_type": "tax_lien",
            "maximum_statutory_return": "110% within 6 months; 115% after 6 months and within 1 year (penalty, not APR)",
            "winning_rate_mechanism": "Highest bid; statutory redemption return is not bid down",
            "redemption_period": "1 year",
            "important_rules": "Buying the certificate does not convey possession or title. Statutory notices and deed steps are required; federal liens and title must be researched independently.",
            "data_source": "Allen County official 2026 Tax Sale List XLSX",
            "last_verified": verified, "source_mode": "live_official_xlsx",
        }
        rows.append(finish(row))
    if len(rows) < 50:
        raise RuntimeError(f"Allen parser found only {len(rows)} rows; refusing suspicious output")
    return rows


def indiana_ad_rows(
    raw: bytes, *, county: str, sale_date: str, sale_time: str,
    sale_location: str, page_url: str, source_url: str, listing_url: str,
    verified: str, expected_prefix: str, minimum_expected: int,
    sale_status: str | None = None,
    important_rules: str | None = None, data_source_label: str | None = None,
) -> list[dict]:
    word_matches = []
    with pdfplumber.open(io.BytesIO(raw)) as pdf:
        text = "\n".join(page.extract_text(layout=True) or "" for page in pdf.pages)
        # Some legal ads contain two text objects that visually overlap.  The
        # high-level text extractor can drop those lines, so also reconstruct
        # each item's header from words sharing the same baseline and column.
        for page in pdf.pages:
            words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
            for head in words:
                if not re.fullmatch(rf"{re.escape(expected_prefix)}\d{{5}}", head["text"]):
                    continue
                line = sorted(
                    (
                        word for word in words
                        if abs(word["top"] - head["top"]) < 1.5
                        and head["x0"] - 1 <= word["x0"] < head["x0"] + 198
                    ),
                    key=lambda word: word["x0"],
                )
                joined = " ".join(word["text"] for word in line)
                found = re.search(
                    rf"\b({re.escape(expected_prefix)}\d{{5}})\s+"
                    r"(\d{2}-\d{2}-\d{2}-\d{3}-\d{3}\.\d{3}-\d{3})\s+"
                    r"\$([\d,]+\.\d{2})\b",
                    joined,
                )
                if found:
                    word_matches.append(found.groups())
    pattern = re.compile(
        rf"\b({re.escape(expected_prefix)}\d{{5}})\s+"
        r"(\d{2}-\d{2}-\d{2}-\d{3}-\d{3}\.\d{3}-\d{3})\s+"
        r"\$([\d,]+\.\d{2})\b"
    )
    # pypdf retains a couple of overlapping headers that pdfplumber may omit.
    alternate_text = "\n".join(page.extract_text() or "" for page in PdfReader(io.BytesIO(raw)).pages)
    matches = pattern.findall(text) + pattern.findall(alternate_text) + word_matches
    dedup = {item: (item, parcel, bid) for item, parcel, bid in matches}
    rows = []
    for item, parcel, bid in dedup.values():
        row = {
            "record_id": f"IN-{county.replace(' ', '')}-2026-{item}",
            "state": "IN", "state_name": "Indiana", "county": county,
            "parcel_id": parcel, "sale_item_number": item,
            "certificate_number": None, "property_address": None, "address": None,
            "city": None, "zip": None, "legal_description": None,
            "property_type": None, "auction_date": sale_date, "sale_date": sale_date,
            "auction_time": sale_time, "auction_format": "In person; county may switch to online",
            "auction_location": sale_location, "auction_url": listing_url,
            "direct_listing_url": listing_url, "official_source_url": source_url,
            "minimum_bid": number(bid), "opening_bid": number(bid),
            "delinquent_tax_amount": None,
            "fees_costs": "Included in minimum bid but not separated per row",
            "assessed_value": None, "market_value": None, "acreage": None,
            "tax_years_delinquent": None,
            "sale_status": sale_status or "Published — subject to removal or change",
            "lien_type": "Indiana tax sale certificate / lien", "sale_type": "tax_lien",
            "maximum_statutory_return": "110% within 6 months; 115% after 6 months and within 1 year (penalty, not APR)",
            "winning_rate_mechanism": "Highest bid; statutory redemption return is not bid down",
            "redemption_period": "Approximately 1 year; confirm certificate-specific deadline",
            "important_rules": important_rules or "The legal advertisement warns that addresses may be inaccurate and the list can change before sale. A certificate is a lien, not immediate ownership.",
            "data_source": data_source_label or f"{county} County official 2026 legal advertisement PDF",
            "last_verified": verified, "source_mode": "live_official_pdf",
            "county_information_url": page_url,
        }
        rows.append(finish(row))
    if len(rows) < minimum_expected:
        raise RuntimeError(f"{county} parser found only {len(rows)} rows; expected at least {minimum_expected}")
    return rows


def coconino_rows(verified: str) -> list[dict]:
    if not LEGACY_PROPERTIES.exists():
        return []
    doc = json.loads(LEGACY_PROPERTIES.read_text(encoding="utf-8"))
    rows = []
    for old in doc.get("properties", []):
        if old.get("state") != "AZ" or old.get("county") != "Coconino" or old.get("sale_type") != "tax_lien":
            continue
        amount = old.get("lien_purchase_amount")
        row = {
            "record_id": f"AZ-Coconino-OTC-{old.get('account_number') or old.get('parcel_id')}",
            "state": "AZ", "state_name": "Arizona", "county": "Coconino",
            "parcel_id": old.get("parcel_id"), "sale_item_number": old.get("case_number"),
            "certificate_number": old.get("account_number"),
            "property_address": old.get("address"), "address": old.get("address"),
            "city": old.get("city"), "zip": old.get("zip"),
            "legal_description": old.get("legal_description"), "property_type": old.get("property_type"),
            "auction_date": None, "sale_date": None, "auction_time": None,
            "auction_format": "Over the counter", "auction_location": "Coconino County Treasurer",
            "auction_url": old.get("auction_url"), "direct_listing_url": old.get("official_url"),
            "official_source_url": old.get("source_document") or old.get("official_url"),
            "minimum_bid": amount, "opening_bid": amount,
            "delinquent_tax_amount": None, "fees_costs": None,
            "assessed_value": old.get("assessed_value"), "market_value": old.get("market_value"),
            "acreage": old.get("acreage"), "tax_years_delinquent": None,
            "sale_status": old.get("sale_status") or "Over-the-counter tax lien available",
            "lien_type": "Arizona over-the-counter tax lien certificate", "sale_type": "tax_lien",
            "maximum_statutory_return": "Rate is certificate-specific; verify with county before purchase",
            "winning_rate_mechanism": "Over-the-counter/state-held certificate; not a live reverse auction",
            "redemption_period": "3 years before foreclosure action may generally begin; verify certificate",
            "important_rules": "A tax-lien certificate is not a deed. The county instructs buyers to research the property and title independently.",
            "data_source": "Coconino County official OTC tax-lien list",
            "last_verified": verified,
            "source_mode": old.get("source_mode") or "verified_official_snapshot",
        }
        rows.append(finish(row))
    return rows


def hamilton_rows(verified: str) -> list[dict]:
    # Live Kendo-grid JSON API behind the county's public tax-sale listing tool.
    # No auth, no paging needed (144 total rows fit in one response). The source
    # publishes an owner-name field (NAME1) alongside each row; it is deliberately
    # never read into a variable here so it cannot leak into published output
    # (see BUG-004/BUG-005 in BUGS.md).
    req = urllib.request.Request(
        HAMILTON_API, data=b"", method="POST",
        headers={
            "User-Agent": USER_AGENT, "Accept": "application/json",
            "Content-Type": "application/x-www-form-urlencoded",
            "X-Requested-With": "XMLHttpRequest",
        },
    )
    context = ssl.create_default_context()
    with urllib.request.urlopen(req, timeout=90, context=context) as response:
        payload = json.loads(response.read())
    rows = []
    for item in payload.get("Data", []):
        sale_id = clean(item.get("Sale_Id"))
        parcel = clean(item.get("Property_Id"))
        if not (sale_id and parcel):
            continue
        address = clean(item.get("Address1"))
        row = {
            "record_id": f"IN-Hamilton-2026-{sale_id}",
            "state": "IN", "state_name": "Indiana", "county": "Hamilton",
            "parcel_id": parcel, "sale_item_number": sale_id,
            "certificate_number": None, "property_address": address, "address": address,
            "city": clean(item.get("LOCATIONCITY")), "zip": clean(item.get("LOCATIONZIP")),
            "legal_description": clean(item.get("Legal_Desc")), "property_type": None,
            "auction_date": "2026-10-08", "sale_date": "2026-10-08",
            "auction_time": "10:00 ET", "auction_format": "In person",
            "auction_location": "Hamilton County Historic Courthouse, 33 N 9th St, "
                                 "2nd Floor Historic Courtroom, Noblesville, IN 46060",
            "auction_url": HAMILTON_LISTING, "direct_listing_url": HAMILTON_LISTING_APP,
            "official_source_url": HAMILTON_LISTING_APP,
            "minimum_bid": number(item.get("MDASCALC")), "opening_bid": number(item.get("MDASCALC")),
            "delinquent_tax_amount": None,
            "fees_costs": "Included in minimum sale price but not separated per row",
            "assessed_value": None, "market_value": None, "acreage": None,
            "tax_years_delinquent": None, "sale_status": "Published — subject to removal or change",
            "lien_type": "Indiana tax sale certificate / lien", "sale_type": "tax_lien",
            "maximum_statutory_return": "110% within 6 months; 115% after 6 months and within 1 year (penalty, not APR)",
            "winning_rate_mechanism": "Highest bid; statutory redemption return is not bid down",
            "redemption_period": "Approximately 1 year; confirm certificate-specific deadline",
            "important_rules": "Official county tax-sale listing tool; the county warns this list is a "
                                "constantly-changing public-records snapshot and may not reflect the most "
                                "current information. A certificate is a lien, not immediate ownership.",
            "data_source": "Hamilton County official 2026 Tax Sale Listing (live JSON grid)",
            "last_verified": verified, "source_mode": "live_official_json_api",
            "county_information_url": HAMILTON_PAGE,
        }
        rows.append(finish(row))
    if len(rows) < 100:
        raise RuntimeError(f"Hamilton parser found only {len(rows)} rows; refusing suspicious output")
    return rows


def prior_by_source() -> dict[str, list[dict]]:
    if not OUTPUT.exists():
        return {}
    try:
        doc = json.loads(OUTPUT.read_text(encoding="utf-8"))
        profiles = doc.get("profiles", {})
        rows = [{**profiles.get(row.get("profile_id"), {}), **row} for row in doc.get("properties", [])]
    except (OSError, json.JSONDecodeError):
        return {}
    result: dict[str, list[dict]] = {}
    for row in rows:
        result.setdefault(row.get("data_source") or "unknown", []).append(row)
    return result


def foreign_entries(managed_profile_ids: set[str]) -> tuple[list[dict], dict[str, dict], list[dict]]:
    """Records already in OUTPUT that belong to a profile this script does not manage.

    This script fully rebuilds the sources it owns on every run, so it must
    never silently drop records another collector wrote for a profile it
    does not manage (e.g. refresh_arizona_cochise_tax_liens.py, which shares
    this same output file). Returns the already-compact property rows and
    their profile entries untouched, so their profile_id and compacted shape
    are preserved exactly as the owning collector wrote them, plus a fully
    reconstructed (profile-merged) copy for accurate summary counts.
    """
    if not OUTPUT.exists():
        return [], {}, []
    try:
        doc = json.loads(OUTPUT.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return [], {}, []
    profiles = doc.get("profiles", {})
    compact = [p for p in doc.get("properties", []) if p.get("profile_id") not in managed_profile_ids]
    kept_profile_ids = {p.get("profile_id") for p in compact}
    kept_profiles = {pid: profiles[pid] for pid in kept_profile_ids if pid in profiles}
    full = [{**profiles.get(p.get("profile_id"), {}), **p} for p in compact]
    return compact, kept_profiles, full


def compact_rows(rows: list[dict]) -> tuple[list[dict], dict[str, dict]]:
    """Deduplicate county-wide metadata while preserving per-row semantics."""
    common_fields = [
        "state", "state_name", "county", "auction_date", "sale_date",
        "auction_time", "auction_format", "auction_location", "auction_url",
        "direct_listing_url", "official_source_url", "fees_costs", "sale_status",
        "lien_type", "sale_type", "maximum_statutory_return",
        "winning_rate_mechanism", "redemption_period", "important_rules",
        "data_source", "last_verified", "source_mode", "county_information_url",
    ]
    groups: dict[str, list[dict]] = {}
    for row in rows:
        profile_id = f"{row.get('state')}-{row.get('county')}-2026"
        row["profile_id"] = profile_id
        groups.setdefault(profile_id, []).append(row)
    profiles: dict[str, dict] = {}
    for profile_id, members in groups.items():
        profile = {}
        for field in common_fields:
            values = [member.get(field) for member in members]
            if values and values[0] is not None and all(value == values[0] for value in values):
                profile[field] = values[0]
        profiles[profile_id] = profile
    compact = []
    for row in rows:
        profile = profiles[row["profile_id"]]
        compact.append({
            key: value for key, value in row.items()
            if value is not None and profile.get(key) != value
        })
    return compact, profiles


def main() -> int:
    now = datetime.now(timezone.utc)
    verified = now.date().isoformat()
    previous = prior_by_source()
    properties: list[dict] = []
    health: list[dict] = []

    sources = [
        (
            "Allen County official 2026 Tax Sale List XLSX", ALLEN_XLSX,
            lambda raw: allen_rows(raw, verified),
        ),
        (
            "Tippecanoe County official 2026 legal advertisement PDF", TIPPECANOE_PDF,
            lambda raw: indiana_ad_rows(
                raw, county="Tippecanoe", sale_date="2026-09-18", sale_time="10:00 ET",
                sale_location="Tippecanoe County Office Building, Tippecanoe Room",
                page_url=TIPPECANOE_PAGE, source_url=TIPPECANOE_PDF,
                listing_url=TIPPECANOE_LISTING, verified=verified,
                expected_prefix="7926", minimum_expected=80,
            ),
        ),
        (
            "Wabash County official 2026 legal advertisement PDF", WABASH_PDF,
            lambda raw: indiana_ad_rows(
                raw, county="Wabash", sale_date="2026-09-03", sale_time=None,
                sale_location="Wabash County Courthouse, Commissioners' Room, 2nd Floor",
                page_url=WABASH_PAGE, source_url=WABASH_PDF,
                listing_url=WABASH_LISTING, verified=verified,
                expected_prefix="8526", minimum_expected=20,
            ),
        ),
        (
            "Grant County official 2026 Commissioners' Certificate Sale advertisement PDF", GRANT_PDF,
            lambda raw: indiana_ad_rows(
                raw, county="Grant", sale_date="2026-04-28", sale_time="10:00 ET",
                sale_location="Grant County Complex, Council Chambers, 1607 S Adams St, Marion, IN 46953",
                page_url=GRANT_PAGE, source_url=GRANT_PDF,
                listing_url=GRANT_LISTING, verified=verified,
                expected_prefix="2725", minimum_expected=50,
                sale_status="Commissioners' Certificate Sale published for 2026-04-28 — that sale date has passed; this is a verified snapshot of the official notice and per-item availability has not been reconfirmed since.",
                important_rules=(
                    "This is a Commissioners' Certificate Sale of certificates the county already holds, distinct from Grant "
                    "County's regular annual tax sale. The advertised sale date has passed as of this refresh; confirm "
                    "current availability, redemption, or purchase status for any specific item with the Grant County "
                    "Auditor before relying on it. A certificate is a lien, not immediate ownership."
                ),
                data_source_label="Grant County official 2026 Commissioners' Certificate Sale advertisement PDF",
            ),
        ),
    ]

    for name, url, parser in sources:
        try:
            rows = parser(fetch(url))
            properties.extend(rows)
            health.append({"source": name, "url": url, "ok": True, "records": len(rows), "checked_at": now.isoformat()})
            print(f"{name}: {len(rows)} rows")
        except Exception as exc:  # preserve verified data on any transient/source format failure
            retained = previous.get(name, [])
            properties.extend(retained)
            health.append({
                "source": name, "url": url, "ok": False, "records": len(retained),
                "checked_at": now.isoformat(), "error": f"{type(exc).__name__}: {exc}",
                "retained_previous_rows": bool(retained),
            })
            print(f"{name}: FAILED ({exc}); retained {len(retained)} prior rows", file=sys.stderr)

    hamilton_name = "Hamilton County official 2026 Tax Sale Listing (live JSON grid)"
    try:
        hamilton = hamilton_rows(verified)
        properties.extend(hamilton)
        health.append({"source": hamilton_name, "url": HAMILTON_API, "ok": True, "records": len(hamilton), "checked_at": now.isoformat()})
        print(f"{hamilton_name}: {len(hamilton)} rows")
    except Exception as exc:  # preserve verified data on any transient/source format failure
        retained = previous.get(hamilton_name, [])
        properties.extend(retained)
        health.append({
            "source": hamilton_name, "url": HAMILTON_API, "ok": False, "records": len(retained),
            "checked_at": now.isoformat(), "error": f"{type(exc).__name__}: {exc}",
            "retained_previous_rows": bool(retained),
        })
        print(f"{hamilton_name}: FAILED ({exc}); retained {len(retained)} prior rows", file=sys.stderr)

    coco = coconino_rows(verified)
    properties.extend(coco)
    health.append({
        "source": "Coconino County official OTC tax-lien list",
        "url": "https://www.coconino.az.gov/376/Tax-Liens", "ok": bool(coco),
        "records": len(coco), "checked_at": now.isoformat(),
        "note": "Normalized from the repository's verified official Coconino snapshot; the county document host returns HTTP 403 to automation.",
    })

    seen = set()
    unique = []
    for row in properties:
        key = row.get("record_id")
        if not key or key in seen:
            continue
        seen.add(key)
        unique.append(row)

    if len(unique) < 100:
        raise SystemExit("Refusing to publish fewer than 100 individual lien records from this script's own sources")

    managed_profile_ids = {f"{row.get('state')}-{row.get('county')}-2026" for row in unique}
    foreign_compact, foreign_profiles, foreign_full = foreign_entries(managed_profile_ids)
    if foreign_full:
        print(f"Preserved {len(foreign_full)} rows from other collectors (e.g. Cochise County) not managed by this script")

    unique.sort(key=lambda row: (row.get("auction_date") or "9999-99-99", row.get("state") or "", row.get("county") or "", row.get("sale_item_number") or ""))
    all_rows_for_counts = unique + foreign_full

    blockers = [
        {
            "source": "SRI statewide property search/download",
            "url": "https://properties.sriservices.com/properties",
            "reason": "The public site is searchable without login, but its host returned a connection-refused/HTTP 502 response to this collector environment. No access control was bypassed.",
        },
        {
            "source": "Kosciusko County 2026 Tax Sale List",
            "url": "https://www.kosciusko.in.gov/egov/apps/document/center.egov?id=4209&view=item",
            "reason": "The county document endpoint returns HTTP 403 to automated retrieval. Search indexing confirms the list exists, but the collector will not bypass the block.",
        },
        {
            "source": "Monroe County 2026 Tax Sale List",
            "url": "https://www.in.gov/counties/monroe/Departments/treasurer/",
            "reason": "The county links the list through SharePoint rather than a stable anonymous CSV/XLSX/PDF download endpoint. No authenticated or cookie-gated collection was attempted.",
        },
    ]

    counts = {
        "total_records": len(all_rows_for_counts),
        "states": len({row.get("state") for row in all_rows_for_counts if row.get("state")}),
        "counties": len({(row.get("state"), row.get("county")) for row in all_rows_for_counts if row.get("state") and row.get("county")}),
        "with_parcel_id": sum(bool(row.get("parcel_id")) for row in all_rows_for_counts),
        "with_address": sum(bool(row.get("property_address")) for row in all_rows_for_counts),
        "with_auction_date": sum(bool(row.get("auction_date")) for row in all_rows_for_counts),
        "with_minimum_bid": sum(row.get("minimum_bid") is not None for row in all_rows_for_counts),
        "with_assessed_value": sum(row.get("assessed_value") is not None for row in all_rows_for_counts),
        "with_research_priority": sum(row.get("research_priority") is not None for row in all_rows_for_counts),
    }
    compact, profiles = compact_rows(unique)
    compact.extend(foreign_compact)
    profiles.update(foreign_profiles)
    output = {
        "schema_version": 1, "updated_at": now.isoformat(), "counts": counts,
        "profiles": profiles, "properties": compact,
        "source_health": health, "blocked_sources": blockers,
        "disclaimer": "Research-priority scores organize follow-up research only and are not recommendations to bid or buy.",
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    # Keep the deploy payload compact; the collector and schema remain readable.
    OUTPUT.write_text(json.dumps(output, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
    print(json.dumps(counts, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
