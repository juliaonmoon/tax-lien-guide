#!/usr/bin/env python3
"""Prefer Linn County's official XLSX publication; fall back to the vetted PDF parser."""
from __future__ import annotations

import io
import re
from datetime import date

import requests
from openpyxl import load_workbook

import refresh_iowa_linn_tax_liens as pdf_linn

SOURCE_XLSX = pdf_linn.SOURCE_XLSX
UA = pdf_linn.UA
PARCEL_RE = re.compile(r"^\d{15}$")


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def fetch_xlsx() -> bytes:
    response = requests.get(
        SOURCE_XLSX,
        headers={"User-Agent": UA, "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*"},
        timeout=90,
    )
    response.raise_for_status()
    if not response.content.startswith(b"PK"):
        raise RuntimeError("Linn County publication did not return an XLSX workbook")
    return response.content


def _parcel(value) -> str | None:
    text = _text(value).replace("-", "").replace(" ", "")
    if PARCEL_RE.fullmatch(text):
        return text
    if text.isdigit() and 13 <= len(text) <= 14:
        candidate = text.zfill(15)
        if PARCEL_RE.fullmatch(candidate):
            return candidate
    return None


def _find_header(rows: list[list[str]]) -> tuple[int, dict[str, int]]:
    """Resolve split/merged spreadsheet headers without touching owner-name data.

    Linn County's CivicPlus workbook can render a logical column heading across
    multiple worksheet rows.  The old parser required "parcel" and "total" to
    appear on the same physical row, which caused a valid workbook to produce
    zero records.  Find the first actual parcel row, then combine only the
    preceding header-band text vertically per column.
    """
    if not rows:
        raise RuntimeError("Linn County XLSX workbook is empty")

    data_start = None
    parcel_candidates: dict[int, int] = {}
    scan_limit = min(len(rows), 160)
    for ridx, row in enumerate(rows[:scan_limit]):
        for col, cell in enumerate(row):
            if _parcel(cell):
                parcel_candidates[col] = parcel_candidates.get(col, 0) + 1
                if data_start is None:
                    data_start = ridx

    if data_start is None or not parcel_candidates:
        raise RuntimeError("Linn County XLSX contains no recognizable parcel rows")

    parcel_col = max(parcel_candidates, key=parcel_candidates.get)
    header_start = max(0, data_start - 30)
    width = max((len(row) for row in rows[header_start:data_start]), default=0)
    combined: list[str] = []
    for col in range(width):
        parts: list[str] = []
        for row in rows[header_start:data_start]:
            if col >= len(row):
                continue
            text = row[col].strip().lower()
            if text and text not in parts:
                parts.append(text)
        combined.append(" ".join(parts))

    columns: dict[str, int] = {"parcel": parcel_col}
    for col, label in enumerate(combined):
        if not label:
            continue
        if "legal" in label or "description" in label:
            columns.setdefault("legal", col)
        if (
            label in {"item", "item #", "item no", "number", "no."}
            or "item number" in label
            or "sale item" in label
        ):
            columns.setdefault("item", col)
        if any(token in label for token in ("total", "amount due", "tax sale amount", "delinquent amount")):
            columns.setdefault("amount", col)

    # Some CivicPlus exports leave the item heading blank/merged. Infer only
    # the item column from the first three columns when its values are a clean
    # 1..1568 sequence. This does not inspect or retain taxpayer/owner names.
    if "item" not in columns:
        best_col = None
        best_hits = 0
        for col in range(min(3, width)):
            hits = 0
            for row in rows[data_start : min(len(rows), data_start + 120)]:
                if col >= len(row):
                    continue
                m = re.fullmatch(r"(\d{1,4})\.?", row[col].strip())
                if m and 1 <= int(m.group(1)) <= 1568:
                    hits += 1
            if hits > best_hits:
                best_col, best_hits = col, hits
        if best_col is not None and best_hits >= 5:
            columns["item"] = best_col

    if "amount" not in columns:
        raise RuntimeError(
            "Linn County XLSX amount/total column was not recognized safely; "
            f"header labels seen: {combined[:20]}"
        )
    return data_start - 1, columns


def parse_xlsx(raw: bytes, verified: str) -> list[dict]:
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    output: list[dict] = []
    seen: set[int] = set()
    public_bidder = False

    for ws in wb.worksheets:
        values = [[_text(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
        try:
            header_idx, cols = _find_header(values)
        except RuntimeError:
            continue

        for row in values[header_idx + 1 :]:
            joined = " ".join(cell for cell in row if cell).lower()
            if "public bidder real estate" in joined:
                public_bidder = True
                continue
            if "mobile home" in joined and ("offer" in joined or "sale" in joined):
                break

            parcel = _parcel(row[cols["parcel"]] if cols["parcel"] < len(row) else "")
            if not parcel:
                continue

            item = None
            if "item" in cols and cols["item"] < len(row):
                m = re.search(r"\d{1,4}", row[cols["item"]])
                if m:
                    item = int(m.group())
            if item is None:
                for cell in row[:3]:
                    m = re.fullmatch(r"(\d{1,4})\.?", cell)
                    if m:
                        item = int(m.group(1))
                        break
            if item is None or not (1 <= item <= 1568) or item in seen:
                continue

            amount_text = row[cols["amount"]] if cols["amount"] < len(row) else ""
            amount_match = re.search(r"-?\$?([\d,]+(?:\.\d{1,2})?)", amount_text)
            if not amount_match:
                continue
            amount = round(float(amount_match.group(1).replace(",", "")), 2)
            if amount < 0:
                continue

            legal = None
            if "legal" in cols and cols["legal"] < len(row):
                legal = re.sub(r"\s+", " ", row[cols["legal"]]).strip() or None

            sale_class = "Public bidder tax sale" if public_bidder else "Regular tax sale"
            redemption = (
                "90-day notice of right of redemption may be issued after 9 months from sale"
                if public_bidder
                else "90-day notice of right of redemption may be issued after 1 year 9 months from sale"
            )
            output.append({
                "record_id": f"IA-Linn-2026-{item}",
                "profile_id": pdf_linn.PROFILE_ID,
                "state": "IA",
                "state_name": "Iowa",
                "county": "Linn",
                "parcel_id": parcel,
                "sale_item_number": str(item),
                "legal_description": legal,
                "auction_date": "2026-06-15",
                "sale_date": "2026-06-15",
                "auction_time": "09:00 CT",
                "auction_format": "Online; percentage-interest bid down with random selection among tied lowest bidders",
                "auction_location": "Online through Iowa Tax Auction; administered by Linn County Treasurer",
                "auction_url": pdf_linn.SOURCE_PAGE,
                "official_source_url": SOURCE_XLSX,
                "direct_listing_url": pdf_linn.SOURCE_PAGE,
                "minimum_bid": None,
                "opening_bid": None,
                "delinquent_tax_amount": amount,
                "sale_status": f"Official 2026 publication snapshot — {sale_class}; sale date has passed and current parcel status must be reconfirmed",
                "lien_type": "Iowa tax sale certificate / property-tax lien",
                "sale_type": "tax_lien",
                "maximum_statutory_return": "2% per month redemption interest; fractions of a month count as a whole month",
                "winning_rate_mechanism": "Parcels are offered at 100%; bidders bid their percentage interest down in whole percentages from 99% to 1%; tied lowest bids are resolved by random selection",
                "redemption_period": redemption,
                "important_rules": "The county publication is a pre-sale delinquent-tax snapshot and may include parcels paid or withheld after publication. A Tax Sale Certificate of Purchase does not convey title; a later Treasurer's Deed process is separate.",
                "data_source": "Linn County Treasurer official 2026 tax-sale publication (Excel)",
                "last_verified": verified,
                "source_mode": "official_2026_publication_snapshot_xlsx",
                "data_completeness": {"published_fields": 8, "tracked_fields": 19, "percent": 42},
                "research_priority": {
                    "score": 42,
                    "label": "Low research priority",
                    "reasons": ["Official parcel ID is published", "Official delinquent-tax amount is published", "The 2026 sale date is confirmed but has passed"],
                    "disclaimer": "Research-priority ranking only; not a buy recommendation.",
                },
            })
            seen.add(item)

    output.sort(key=lambda row: int(row["sale_item_number"]))
    if len(output) < 1500:
        raise RuntimeError(f"Linn County XLSX parser found only {len(output)} rows; using PDF fallback")
    if any(int(row["sale_item_number"]) > 1568 for row in output):
        raise RuntimeError("Linn County XLSX parser crossed into the mobile-home section")
    return output


def main() -> None:
    try:
        rows = parse_xlsx(fetch_xlsx(), date.today().isoformat())
    except (requests.RequestException, RuntimeError, ValueError) as exc:
        print(f"Linn County IA: Excel-first refresh unavailable/unrecognized; falling back to vetted PDF parser. Reason: {exc}")
        pdf_linn.main()
        return
    pdf_linn.update_details(rows)
    print(f"Linn County IA: loaded {len(rows)} official XLSX real-estate tax-lien rows; owner names intentionally omitted")


if __name__ == "__main__":
    main()
